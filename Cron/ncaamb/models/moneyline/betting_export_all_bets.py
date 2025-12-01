#!/usr/bin/env python3
"""
Export ALL 2025 bets - both team_1 and team_2 for each game
Shows what would have happened if you bet on either team
"""

import polars as pl
import lightgbm as lgb
import xgboost as xgb
import numpy as np
import os
import pandas as pd
from datetime import datetime

def load_features(years):
    """Load and combine features from multiple years"""
    dfs = []
    features_dir = os.path.join(os.path.dirname(__file__), '..', '..')

    first_loaded = False
    reference_schema = None

    for year in years:
        file = os.path.join(features_dir, f"features{year}.csv")
        if not os.path.exists(file):
            continue
        print(f"[+] Loading {file}...")
        df = pl.read_csv(file)

        if not first_loaded:
            for col in df.columns:
                if df[col].dtype in [pl.Float32, pl.Float64]:
                    df = df.with_columns(pl.col(col).cast(pl.Float64))
                elif df[col].dtype in [pl.Int32, pl.Int64]:
                    df = df.with_columns(pl.col(col).cast(pl.Int64))
            reference_schema = df.schema
            first_loaded = True
        else:
            for col in df.columns:
                if col in reference_schema:
                    target_dtype = reference_schema[col]
                    if df[col].dtype != target_dtype:
                        try:
                            df = df.with_columns(pl.col(col).cast(target_dtype))
                        except:
                            pass

        dfs.append(df)

    combined = pl.concat(dfs, how='diagonal')
    return combined

def create_target_variable(df):
    df = df.with_columns([
        pl.when(
            (pl.col('team_1_score').is_not_null()) &
            (pl.col('team_2_score').is_not_null())
        ).then(
            pl.when(pl.col('team_1_score') > pl.col('team_2_score')).then(1).otherwise(0)
        ).otherwise(None).alias('ml_target')
    ])
    return df

def preprocess_features(df):
    odds_cols = [
        'betonline_ml_team_1', 'betonline_ml_team_2',
        'betonline_spread_odds_team_1', 'betonline_spread_odds_team_2',
        'betonline_spread_pts_team_1', 'betonline_spread_pts_team_2',
        'fanduel_spread_odds_team_1', 'fanduel_spread_odds_team_2',
        'fanduel_spread_pts_team_1', 'fanduel_spread_pts_team_2',
        'mybookie_ml_team_1', 'mybookie_ml_team_2',
        'mybookie_spread_odds_team_1', 'mybookie_spread_odds_team_2',
        'mybookie_spread_pts_team_1', 'mybookie_spread_pts_team_2',
    ]
    for col in odds_cols:
        if col in df.columns:
            df = df.with_columns(pl.col(col).cast(pl.Float64))
    return df

def filter_quality_games(df):
    initial_count = len(df)
    df = df.filter(
        (pl.col('team_1_data_quality') >= 0.5) &
        (pl.col('team_2_data_quality') >= 0.5)
    )
    df = df.filter(pl.col('ml_target').is_not_null())
    df = df.filter(
        (pl.col('avg_ml_team_1').is_not_null()) &
        (pl.col('avg_ml_team_2').is_not_null())
    )
    removed = initial_count - len(df)
    print(f"[+] Filtered: {initial_count} -> {len(df)} ({removed} removed)")
    return df

def get_feature_columns(df):
    metadata_cols = {
        'game_id', 'date', 'season', 'team_1', 'team_2',
        'team_1_score', 'team_2_score', 'actual_total',
        'team_1_conference', 'team_2_conference',
        'team_1_is_home', 'team_2_is_home', 'location',
        'total_score_outcome', 'team_1_winloss',
        'team_1_leaderboard', 'team_2_leaderboard',
        'team_1_match_hist', 'team_2_match_hist',
        'team_1_hist_count', 'team_2_hist_count',
        'start_time', 'game_odds', 'ml_target',
        'team_1_data_quality', 'team_2_data_quality'
    }
    feature_cols = []
    for col in df.columns:
        if col not in metadata_cols:
            if df[col].dtype in [pl.Float32, pl.Float64, pl.Int32, pl.Int64]:
                feature_cols.append(col)
    return feature_cols

def prepare_training_data(df, feature_cols):
    df = df.with_columns([
        pl.col(col).fill_null(0) for col in feature_cols
    ])
    X = df.select(feature_cols).to_numpy()
    y = df.select('ml_target').to_numpy().ravel()
    return X, y

def american_to_decimal(american_odds):
    """Convert American odds to decimal"""
    american_odds = float(american_odds)
    if american_odds >= 0:
        return (american_odds / 100) + 1
    else:
        return (100 / abs(american_odds)) + 1

def calculate_ev(win_prob, decimal_odds):
    """Calculate Expected Value"""
    win_prob = float(win_prob)
    decimal_odds = float(decimal_odds)
    ev = (win_prob * (decimal_odds - 1)) - (1 - win_prob)
    return ev

def get_best_odds(row, team):
    """Get best odds for team, preferring Bovada, then Betonline, then FanDuel"""
    if team == 'team_1':
        # Prefer Bovada
        if row.get('bovada_ml_team_1') and row.get('bovada_ml_team_1') != 0:
            return float(row['bovada_ml_team_1']), 'Bovada'
        # Fallback to Betonline
        if row.get('betonline_ml_team_1') and row.get('betonline_ml_team_1') != 0:
            return float(row['betonline_ml_team_1']), 'Betonline'
        # Fallback to FanDuel
        if row.get('fanduel_ml_team_1') and row.get('fanduel_ml_team_1') != 0:
            return float(row['fanduel_ml_team_1']), 'FanDuel'
        # Last resort: average
        if row.get('avg_ml_team_1') and row.get('avg_ml_team_1') != 0:
            return float(row['avg_ml_team_1']), 'Average'
    else:  # team_2
        # Prefer Bovada
        if row.get('bovada_ml_team_2') and row.get('bovada_ml_team_2') != 0:
            return float(row['bovada_ml_team_2']), 'Bovada'
        # Fallback to Betonline
        if row.get('betonline_ml_team_2') and row.get('betonline_ml_team_2') != 0:
            return float(row['betonline_ml_team_2']), 'Betonline'
        # Fallback to FanDuel
        if row.get('fanduel_ml_team_2') and row.get('fanduel_ml_team_2') != 0:
            return float(row['fanduel_ml_team_2']), 'FanDuel'
        # Last resort: average
        if row.get('avg_ml_team_2') and row.get('avg_ml_team_2') != 0:
            return float(row['avg_ml_team_2']), 'Average'

    return None, None

def main():
    print("="*80)
    print("EXPORT ALL 2025 BETS - BOTH TEAMS PER GAME")
    print("="*80)

    # Load data
    print("\n[STEP 1] Loading training data (2021-2024)...")
    train_df = load_features(['2021', '2022', '2023', '2024'])

    print("[STEP 2] Loading test data (2025)...")
    test_df = load_features(['2025'])

    # Prepare
    print("[STEP 3] Processing data...")
    train_df = create_target_variable(train_df)
    test_df = create_target_variable(test_df)
    train_df = preprocess_features(train_df)
    test_df = preprocess_features(test_df)

    print("[STEP 4] Filtering quality games...")
    train_df = filter_quality_games(train_df)
    test_df = filter_quality_games(test_df)

    feature_cols = get_feature_columns(train_df)
    print(f"[+] Features: {len(feature_cols)}")

    print("[STEP 5] Preparing data...")
    X_train, y_train = prepare_training_data(train_df, feature_cols)
    X_test, y_test = prepare_training_data(test_df, feature_cols)

    print(f"[+] Train: {X_train.shape[0]} games, Test: {X_test.shape[0]} games")

    # Train models
    print("\n[STEP 6] Training LightGBM...")
    lgb_params = {
        'num_leaves': 10,
        'learning_rate': 0.011905546738777037,
        'feature_fraction': 0.5,
        'bagging_fraction': 0.6902301680678105,
        'min_data_in_leaf': 100,
        'lambda_l1': 5,
        'lambda_l2': 5,
        'bagging_freq': 5,
        'objective': 'binary',
        'metric': 'binary_logloss',
        'verbose': -1,
        'seed': 42,
    }

    train_data = lgb.Dataset(X_train, label=y_train)
    test_data = lgb.Dataset(X_test, label=y_test, reference=train_data)
    lgb_model = lgb.train(
        lgb_params,
        train_data,
        num_boost_round=200,
        valid_sets=[train_data, test_data],
        callbacks=[lgb.log_evaluation(period=100), lgb.early_stopping(50)],
    )

    print("[STEP 7] Training XGBoost...")
    xgb_params = {
        'max_depth': 5,
        'learning_rate': 0.186544,
        'subsample': 0.9937,
        'colsample_bytree': 0.4005,
        'min_child_weight': 15,
        'gamma': 9.7163,
        'lambda': 10.1068,
        'n_estimators': 79,
        'objective': 'binary:logistic',
        'eval_metric': 'logloss',
        'random_state': 42,
        'verbosity': 0,
    }

    xgb_model = xgb.XGBClassifier(**xgb_params)
    xgb_model.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=False)

    # Get predictions
    print("\n[STEP 8] Getting predictions...")
    lgb_pred = lgb_model.predict(X_test)
    xgb_pred = xgb_model.predict_proba(X_test)[:, 1]

    # Ensemble with optimal weighting (LGB 18% + XGB 82%)
    ensemble_pred = (lgb_pred * 0.18) + (xgb_pred * 0.82)

    # Convert test_df to dict for easier access
    test_dict = test_df.to_dicts()

    print("\n[STEP 9] Creating betting records for all games...")
    all_bets = []

    for i, game in enumerate(test_dict):
        game_id = game.get('game_id', '')
        date = game.get('date', '')
        team_1 = game.get('team_1', '')
        team_2 = game.get('team_2', '')
        actual = y_test[i]

        # Ensemble predictions
        prob_team_1 = ensemble_pred[i]
        prob_team_2 = 1 - prob_team_1

        # ===== BET ON TEAM_1 =====
        odds_t1, book_t1 = get_best_odds(game, 'team_1')
        if odds_t1 is not None:
            decimal_t1 = american_to_decimal(odds_t1)
            ev_t1 = calculate_ev(prob_team_1, decimal_t1)

            # Outcome: did team_1 win?
            team_1_won = (actual == 1)
            if team_1_won:
                pnl_t1 = 10 * (decimal_t1 - 1)
                outcome_t1 = 'Win'
            else:
                pnl_t1 = -10
                outcome_t1 = 'Loss'

            all_bets.append({
                'game_id': game_id,
                'date': date,
                'team_1': team_1,
                'team_2': team_2,
                'bet_team': team_1,
                'prob_bet_team': round(prob_team_1, 4),
                'sportsbook': book_t1,
                'odds_american': round(odds_t1, 0),
                'odds_decimal': round(decimal_t1, 3),
                'ev_%': round(ev_t1 * 100, 2),
                'outcome': outcome_t1,
                'pnl': round(pnl_t1, 2),
            })

        # ===== BET ON TEAM_2 =====
        odds_t2, book_t2 = get_best_odds(game, 'team_2')
        if odds_t2 is not None:
            decimal_t2 = american_to_decimal(odds_t2)
            ev_t2 = calculate_ev(prob_team_2, decimal_t2)

            # Outcome: did team_2 win?
            team_2_won = (actual == 0)
            if team_2_won:
                pnl_t2 = 10 * (decimal_t2 - 1)
                outcome_t2 = 'Win'
            else:
                pnl_t2 = -10
                outcome_t2 = 'Loss'

            all_bets.append({
                'game_id': game_id,
                'date': date,
                'team_1': team_1,
                'team_2': team_2,
                'bet_team': team_2,
                'prob_bet_team': round(prob_team_2, 4),
                'sportsbook': book_t2,
                'odds_american': round(odds_t2, 0),
                'odds_decimal': round(decimal_t2, 3),
                'ev_%': round(ev_t2 * 100, 2),
                'outcome': outcome_t2,
                'pnl': round(pnl_t2, 2),
            })

    # Export to Excel
    df_bets = pd.DataFrame(all_bets)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = os.path.join(os.path.dirname(__file__), 'saved', f'all_bets_2025_{timestamp}.xlsx')
    os.makedirs(os.path.dirname(excel_filename), exist_ok=True)

    # Write to Excel with formatting
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df_bets.to_excel(writer, sheet_name='All_Bets', index=False)

        # Format the workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = writer.book
        worksheet = writer.sheets['All_Bets']

        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color code wins (green) and losses (red)
        for row in worksheet.iter_rows(min_row=2, max_row=len(df_bets)+1):
            for cell in row:
                if cell.column == 12:  # PnL column
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        for col, width in [('A', 18), ('B', 12), ('C', 20), ('D', 20), ('E', 18),
                           ('F', 14), ('G', 12), ('H', 14), ('I', 14), ('J', 10),
                           ('K', 10), ('L', 10)]:
            worksheet.column_dimensions[col].width = width

    print(f"\n[+] Excel file saved: {excel_filename}")
    print(f"[+] Total bets: {len(all_bets)} (2 bets per game)")
    print(f"[+] Games: {len(test_dict)}")

    # Summary stats
    wins = len([b for b in all_bets if b['outcome'] == 'Win'])
    losses = len([b for b in all_bets if b['outcome'] == 'Loss'])
    total_pnl = sum(b['pnl'] for b in all_bets)
    win_pct = (wins / len(all_bets) * 100) if len(all_bets) > 0 else 0
    roi = (total_pnl / (len(all_bets) * 10) * 100) if len(all_bets) > 0 else 0

    print(f"\n[SUMMARY - ALL BETS]")
    print(f"  Total Bets: {len(all_bets)}")
    print(f"  Wins: {wins} ({win_pct:.1f}%)")
    print(f"  Losses: {losses}")
    print(f"  Total PnL: ${total_pnl:,.2f}")
    print(f"  ROI: {roi:.2f}%")
    print("="*80)

if __name__ == "__main__":
    main()
