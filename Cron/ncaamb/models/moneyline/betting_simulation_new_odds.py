#!/usr/bin/env python3
"""
Comprehensive Moneyline Betting Simulation
EV-based and Confidence-based betting with optimal ensemble weighting (LGB 18% + XGB 82%)
Train 2021-2024, test 2025
"""

import polars as pl
import lightgbm as lgb
import xgboost as xgb
import numpy as np
import os
from sklearn.metrics import accuracy_score
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def load_features(years):
    """Load and combine features from multiple years"""
    dfs = []
    features_dir = os.path.join(os.path.dirname(__file__), '..', '..')

    first_loaded = False
    reference_schema = None

    for year in years:
        file = os.path.join(features_dir, f"features{year}.csv")
        if not os.path.exists(file):
            continue
        print(f"[+] Loading {file}...")
        df = pl.read_csv(file)

        if not first_loaded:
            for col in df.columns:
                if df[col].dtype in [pl.Float32, pl.Float64]:
                    df = df.with_columns(pl.col(col).cast(pl.Float64))
                elif df[col].dtype in [pl.Int32, pl.Int64]:
                    df = df.with_columns(pl.col(col).cast(pl.Int64))
            reference_schema = df.schema
            first_loaded = True
        else:
            for col in df.columns:
                if col in reference_schema:
                    target_dtype = reference_schema[col]
                    if df[col].dtype != target_dtype:
                        try:
                            df = df.with_columns(pl.col(col).cast(target_dtype))
                        except:
                            pass

        dfs.append(df)

    combined = pl.concat(dfs, how='diagonal')
    return combined

def create_target_variable(df):
    df = df.with_columns([
        pl.when(
            (pl.col('team_1_score').is_not_null()) &
            (pl.col('team_2_score').is_not_null())
        ).then(
            pl.when(pl.col('team_1_score') > pl.col('team_2_score')).then(1).otherwise(0)
        ).otherwise(None).alias('ml_target')
    ])
    return df

def preprocess_features(df):
    odds_cols = [
        'betonline_ml_team_1', 'betonline_ml_team_2',
        'betonline_spread_odds_team_1', 'betonline_spread_odds_team_2',
        'betonline_spread_pts_team_1', 'betonline_spread_pts_team_2',
        'fanduel_spread_odds_team_1', 'fanduel_spread_odds_team_2',
        'fanduel_spread_pts_team_1', 'fanduel_spread_pts_team_2',
        'mybookie_ml_team_1', 'mybookie_ml_team_2',
        'mybookie_spread_odds_team_1', 'mybookie_spread_odds_team_2',
        'mybookie_spread_pts_team_1', 'mybookie_spread_pts_team_2',
    ]
    for col in odds_cols:
        if col in df.columns:
            df = df.with_columns(pl.col(col).cast(pl.Float64))
    return df

def filter_quality_games(df):
    initial_count = len(df)
    df = df.filter(
        (pl.col('team_1_data_quality') >= 0.5) &
        (pl.col('team_2_data_quality') >= 0.5)
    )
    df = df.filter(pl.col('ml_target').is_not_null())
    df = df.filter(
        (pl.col('avg_ml_team_1').is_not_null()) &
        (pl.col('avg_ml_team_2').is_not_null())
    )
    removed = initial_count - len(df)
    print(f"[+] Filtered: {initial_count} -> {len(df)} ({removed} removed)")
    return df

def get_feature_columns(df):
    metadata_cols = {
        'game_id', 'date', 'season', 'team_1', 'team_2',
        'team_1_score', 'team_2_score', 'actual_total',
        'team_1_conference', 'team_2_conference',
        'team_1_is_home', 'team_2_is_home', 'location',
        'total_score_outcome', 'team_1_winloss',
        'team_1_leaderboard', 'team_2_leaderboard',
        'team_1_match_hist', 'team_2_match_hist',
        'team_1_hist_count', 'team_2_hist_count',
        'start_time', 'game_odds', 'ml_target',
        'team_1_data_quality', 'team_2_data_quality'
    }
    feature_cols = []
    for col in df.columns:
        if col not in metadata_cols:
            if df[col].dtype in [pl.Float32, pl.Float64, pl.Int32, pl.Int64]:
                feature_cols.append(col)
    return feature_cols

def prepare_training_data(df, feature_cols):
    df = df.with_columns([
        pl.col(col).fill_null(0) for col in feature_cols
    ])
    X = df.select(feature_cols).to_numpy()
    y = df.select('ml_target').to_numpy().ravel()
    return X, y

def american_to_decimal(american_odds):
    """Convert American odds to decimal"""
    american_odds = float(american_odds)
    if american_odds >= 0:
        return (american_odds / 100) + 1
    else:
        return (100 / abs(american_odds)) + 1

def calculate_ev(win_prob, decimal_odds):
    """Calculate Expected Value"""
    win_prob = float(win_prob)
    decimal_odds = float(decimal_odds)
    ev = (win_prob * (decimal_odds - 1)) - (1 - win_prob)
    return ev

def get_best_odds(row):
    """Get best moneyline odds, preferring Bovada, then Betonline"""
    # Prefer Bovada first
    if row.get('bovada_ml_team_1') and row.get('bovada_ml_team_1') != 0:
        return float(row['bovada_ml_team_1']), 'Bovada'

    # Fallback to Betonline
    if row.get('betonline_ml_team_1') and row.get('betonline_ml_team_1') != 0:
        return float(row['betonline_ml_team_1']), 'Betonline'

    # If neither available, use average
    if row.get('avg_ml_team_1') and row.get('avg_ml_team_1') != 0:
        return float(row['avg_ml_team_1']), 'Average'

    return None, None

def main():
    print("="*80)
    print("MONEYLINE BETTING SIMULATION - NEW DECIMAL ODDS")
    print("Ensemble: LGB 18% + XGB 82%")
    print("="*80)

    # Load data
    print("\n[STEP 1] Loading training data (2021-2024)...")
    train_df = load_features(['2021', '2022', '2023', '2024'])

    print("[STEP 2] Loading test data (2025)...")
    test_df = load_features(['2025'])

    # Prepare
    print("[STEP 3] Processing data...")
    train_df = create_target_variable(train_df)
    test_df = create_target_variable(test_df)
    train_df = preprocess_features(train_df)
    test_df = preprocess_features(test_df)

    print("[STEP 4] Filtering quality games...")
    train_df = filter_quality_games(train_df)
    test_df = filter_quality_games(test_df)

    feature_cols = get_feature_columns(train_df)
    print(f"[+] Features: {len(feature_cols)}")

    print("[STEP 5] Preparing data...")
    X_train, y_train = prepare_training_data(train_df, feature_cols)
    X_test, y_test = prepare_training_data(test_df, feature_cols)

    print(f"[+] Train: {X_train.shape[0]} games, Test: {X_test.shape[0]} games")

    # Train models
    print("\n[STEP 6] Training LightGBM...")
    lgb_params = {
        'num_leaves': 10,
        'learning_rate': 0.011905546738777037,
        'feature_fraction': 0.5,
        'bagging_fraction': 0.6902301680678105,
        'min_data_in_leaf': 100,
        'lambda_l1': 5,
        'lambda_l2': 5,
        'bagging_freq': 5,
        'objective': 'binary',
        'metric': 'binary_logloss',
        'verbose': -1,
        'seed': 42,
    }

    train_data = lgb.Dataset(X_train, label=y_train)
    test_data = lgb.Dataset(X_test, label=y_test, reference=train_data)
    lgb_model = lgb.train(
        lgb_params,
        train_data,
        num_boost_round=200,
        valid_sets=[train_data, test_data],
        callbacks=[lgb.log_evaluation(period=100), lgb.early_stopping(50)],
    )

    print("[STEP 7] Training XGBoost...")
    xgb_params = {
        'max_depth': 5,
        'learning_rate': 0.186544,
        'subsample': 0.9937,
        'colsample_bytree': 0.4005,
        'min_child_weight': 15,
        'gamma': 9.7163,
        'lambda': 10.1068,
        'n_estimators': 79,
        'objective': 'binary:logistic',
        'eval_metric': 'logloss',
        'random_state': 42,
        'verbosity': 0,
    }

    xgb_model = xgb.XGBClassifier(**xgb_params)
    xgb_model.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=False)

    # Get predictions
    print("\n[STEP 8] Getting predictions...")
    lgb_pred = lgb_model.predict(X_test)
    xgb_pred = xgb_model.predict_proba(X_test)[:, 1]

    # Ensemble with optimal weighting (LGB 18% + XGB 82%)
    ensemble_pred = (lgb_pred * 0.18) + (xgb_pred * 0.82)

    # Convert test_df to dict for easier access
    test_dict = test_df.to_dicts()

    # Get moneyline odds and actual results
    ml_odds = []
    best_books = []
    actuals = []

    for i, game in enumerate(test_dict):
        best_odds, best_book = get_best_odds(game)
        if best_odds is None:
            best_odds = game.get('avg_ml_team_1', 0)
            best_book = 'avg'

        ml_odds.append(best_odds)
        best_books.append(best_book)
        actuals.append(y_test[i])

    ml_odds = np.array(ml_odds)
    actuals = np.array(actuals, dtype=int)

    # =========================================================================
    # SIMULATION 1: EV-BASED BETTING
    # =========================================================================
    print("\n" + "="*80)
    print("SIMULATION 1: EV-BASED MONEYLINE BETTING")
    print("="*80)

    # Calculate EV for each game
    decimal_odds = np.array([american_to_decimal(o) for o in ml_odds])
    evs = np.array([calculate_ev(ensemble_pred[i], decimal_odds[i]) for i in range(len(ensemble_pred))])

    # EV-based simulation with detailed game data
    ev_results = {}
    ev_game_data = {}  # Store detailed game data per threshold
    ev_thresholds = list(range(-5, int(np.max(evs * 100)) + 5))

    for threshold in ev_thresholds:
        threshold_decimal = threshold / 100.0
        mask = evs >= threshold_decimal

        if np.sum(mask) == 0:
            continue

        bets = []
        game_data = []

        for idx in np.where(mask)[0]:
            game = test_dict[idx]
            pred = ensemble_pred[idx]
            actual = actuals[idx]
            decimal = decimal_odds[idx]
            american = ml_odds[idx]
            book = best_books[idx]

            # Determine which team to bet on
            if pred > 0.5:
                bet_team = game.get('team_1', 'team_1')
            else:
                bet_team = game.get('team_2', 'team_2')

            correct = (pred > 0.5) == (actual == 1)

            if correct:
                pnl = 10 * (decimal - 1)
                outcome = 'Win'
            else:
                pnl = -10
                outcome = 'Loss'

            game_data.append({
                'game_id': game.get('game_id', ''),
                'date': game.get('date', ''),
                'team_1': game.get('team_1', ''),
                'team_2': game.get('team_2', ''),
                'bet_team': bet_team,
                'prob_team_1': round(pred, 4),
                'prob_team_2': round(1 - pred, 4),
                'sportsbook': book,
                'odds_american': american,
                'odds_decimal': round(decimal, 3),
                'outcome': outcome,
                'pnl': round(pnl, 2),
            })

            bets.append({'correct': correct, 'pnl': pnl})

        total_pnl = sum(b['pnl'] for b in bets)
        wins = sum(b['correct'] for b in bets)
        losses = len(bets) - wins
        win_pct = (wins / len(bets) * 100) if len(bets) > 0 else 0
        roi = (total_pnl / (len(bets) * 10) * 100) if len(bets) > 0 else 0

        ev_results[threshold] = {
            'threshold': threshold,
            'num_bets': len(bets),
            'wins': wins,
            'losses': losses,
            'win_pct': win_pct,
            'net_profit': total_pnl,
            'roi': roi,
        }
        ev_game_data[threshold] = game_data

    # Print EV results
    print(f"\n{'EV Threshold':<15} {'Bets':<8} {'Wins':<8} {'Win %':<10} {'Net Profit':<15} {'ROI':<10}")
    print("-"*80)

    positive_ev_thresholds = [t for t in sorted(ev_results.keys()) if ev_results[t]['roi'] > 0]

    if positive_ev_thresholds:
        for threshold in positive_ev_thresholds:
            s = ev_results[threshold]
            print(f"{threshold:>14}% {s['num_bets']:<8} {s['wins']:<8} {s['win_pct']:<10.1f}% "
                  f"${s['net_profit']:<14.2f} {s['roi']:<9.2f}%")

    # Find best
    if positive_ev_thresholds:
        best_ev_threshold = max(positive_ev_thresholds, key=lambda t: ev_results[t]['roi'])
        best_ev_stats = ev_results[best_ev_threshold]
        print(f"\nBest EV Threshold: {best_ev_threshold}%")
        print(f"  Bets: {best_ev_stats['num_bets']}")
        print(f"  Wins: {best_ev_stats['wins']} ({best_ev_stats['win_pct']:.1f}%)")
        print(f"  Net Profit: ${best_ev_stats['net_profit']:.2f}")
        print(f"  ROI: {best_ev_stats['roi']:.2f}%")

    # =========================================================================
    # SIMULATION 2: CONFIDENCE-BASED BETTING
    # =========================================================================
    print("\n" + "="*80)
    print("SIMULATION 2: CONFIDENCE-BASED MONEYLINE BETTING")
    print("="*80)
    print("(Confidence = Probability team_1 wins)")
    print("(If confidence < 0.5, bet on team_2 in that confidence range)\n")

    # Create confidence buckets with detailed game data
    conf_results = {}
    conf_game_data = {}  # Store detailed game data per threshold
    conf_steps = np.arange(0.51, 1.01, 0.01)

    for conf_threshold in conf_steps:
        # Identify games where confidence >= threshold OR confidence <= (1 - threshold)
        team1_bets = ensemble_pred >= conf_threshold
        team2_bets = ensemble_pred <= (1 - conf_threshold)
        mask = team1_bets | team2_bets

        if np.sum(mask) == 0:
            continue

        bets = []
        game_data = []

        for idx in np.where(mask)[0]:
            game = test_dict[idx]
            pred = ensemble_pred[idx]
            actual = actuals[idx]
            decimal = decimal_odds[idx]
            american = ml_odds[idx]
            book = best_books[idx]

            # Determine which team we're betting on
            if pred >= conf_threshold:
                # Betting on team_1
                bet_team = game.get('team_1', 'team_1')
                correct = (actual == 1)
            else:
                # Betting on team_2
                bet_team = game.get('team_2', 'team_2')
                correct = (actual == 0)

            if correct:
                pnl = 10 * (decimal - 1)
                outcome = 'Win'
            else:
                pnl = -10
                outcome = 'Loss'

            game_data.append({
                'game_id': game.get('game_id', ''),
                'date': game.get('date', ''),
                'team_1': game.get('team_1', ''),
                'team_2': game.get('team_2', ''),
                'bet_team': bet_team,
                'prob_team_1': round(pred, 4),
                'prob_team_2': round(1 - pred, 4),
                'sportsbook': book,
                'odds_american': american,
                'odds_decimal': round(decimal, 3),
                'outcome': outcome,
                'pnl': round(pnl, 2),
            })

            bets.append({'correct': correct, 'pnl': pnl})

        if len(bets) == 0:
            continue

        total_pnl = sum(b['pnl'] for b in bets)
        wins = sum(b['correct'] for b in bets)
        losses = len(bets) - wins
        win_pct = (wins / len(bets) * 100) if len(bets) > 0 else 0
        roi = (total_pnl / (len(bets) * 10) * 100) if len(bets) > 0 else 0

        conf_results[conf_threshold] = {
            'threshold': conf_threshold,
            'num_bets': len(bets),
            'wins': wins,
            'losses': losses,
            'win_pct': win_pct,
            'net_profit': total_pnl,
            'roi': roi,
        }
        conf_game_data[conf_threshold] = game_data

    # Print confidence results
    print(f"{'Confidence':<15} {'Bets':<8} {'Wins':<8} {'Win %':<10} {'Net Profit':<15} {'ROI':<10}")
    print("-"*80)

    positive_conf_thresholds = [t for t in sorted(conf_results.keys()) if conf_results[t]['roi'] > 0]

    if positive_conf_thresholds:
        for threshold in positive_conf_thresholds:
            s = conf_results[threshold]
            print(f"{threshold:.2f}        {s['num_bets']:<8} {s['wins']:<8} {s['win_pct']:<10.1f}% "
                  f"${s['net_profit']:<14.2f} {s['roi']:<9.2f}%")

    # Find best
    if positive_conf_thresholds:
        best_conf_threshold = max(positive_conf_thresholds, key=lambda t: conf_results[t]['roi'])
        best_conf_stats = conf_results[best_conf_threshold]
        print(f"\nBest Confidence Threshold: {best_conf_threshold:.2f}")
        print(f"  Bets: {best_conf_stats['num_bets']}")
        print(f"  Wins: {best_conf_stats['wins']} ({best_conf_stats['win_pct']:.1f}%)")
        print(f"  Net Profit: ${best_conf_stats['net_profit']:.2f}")
        print(f"  ROI: {best_conf_stats['roi']:.2f}%")

    # =========================================================================
    # EXPORT TO EXCEL
    # =========================================================================
    print("\n" + "="*80)
    print("EXPORTING DETAILED BETTING DATA TO EXCEL")
    print("="*80)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = os.path.join(os.path.dirname(__file__), 'saved', f'betting_simulation_{timestamp}.xlsx')
    os.makedirs(os.path.dirname(excel_filename), exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # --------- EV-BASED SHEET ---------
    # Summary sheet for EV
    ws_ev_summary = wb.create_sheet('EV_Summary')
    ev_summary_data = []
    for threshold in sorted(ev_results.keys()):
        s = ev_results[threshold]
        ev_summary_data.append({
            'EV_Threshold_%': threshold,
            'Num_Bets': s['num_bets'],
            'Wins': s['wins'],
            'Losses': s['losses'],
            'Win_%': round(s['win_pct'], 2),
            'Net_Profit': round(s['net_profit'], 2),
            'ROI_%': round(s['roi'], 2),
        })

    ev_summary_df = pd.DataFrame(ev_summary_data)
    for r_idx, row in enumerate(dataframe_to_rows(ev_summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_ev_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Detailed sheets for each EV threshold
    for threshold in sorted(ev_game_data.keys()):
        ws = wb.create_sheet(f'EV_{threshold}pct')
        stats = ev_results[threshold]

        # Add summary header
        ws['A1'] = f'EV Threshold: {threshold}%'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f'Total Bets: {stats["num_bets"]} | Wins: {stats["wins"]} | Losses: {stats["losses"]} | Win %: {stats["win_pct"]:.1f}% | Net Profit: ${stats["net_profit"]:.2f} | ROI: {stats["roi"]:.2f}%'
        ws['A2'].font = Font(italic=True)

        # Add game data
        game_df = pd.DataFrame(ev_game_data[threshold])
        for r_idx, row in enumerate(dataframe_to_rows(game_df, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 4:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif isinstance(value, (int, float)) and c_idx >= 8:  # PnL columns
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

    # --------- CONFIDENCE-BASED SHEET ---------
    # Summary sheet for Confidence
    ws_conf_summary = wb.create_sheet('Confidence_Summary')
    conf_summary_data = []
    for threshold in sorted(conf_results.keys()):
        s = conf_results[threshold]
        conf_summary_data.append({
            'Confidence_Threshold': round(threshold, 2),
            'Num_Bets': s['num_bets'],
            'Wins': s['wins'],
            'Losses': s['losses'],
            'Win_%': round(s['win_pct'], 2),
            'Net_Profit': round(s['net_profit'], 2),
            'ROI_%': round(s['roi'], 2),
        })

    conf_summary_df = pd.DataFrame(conf_summary_data)
    for r_idx, row in enumerate(dataframe_to_rows(conf_summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_conf_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Detailed sheets for each Confidence threshold
    for threshold in sorted(conf_game_data.keys()):
        threshold_str = f"{threshold:.2f}".replace('.', '_')
        ws = wb.create_sheet(f'Conf_{threshold_str}')
        stats = conf_results[threshold]

        # Add summary header
        ws['A1'] = f'Confidence Threshold: {threshold:.2f}'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f'Total Bets: {stats["num_bets"]} | Wins: {stats["wins"]} | Losses: {stats["losses"]} | Win %: {stats["win_pct"]:.1f}% | Net Profit: ${stats["net_profit"]:.2f} | ROI: {stats["roi"]:.2f}%'
        ws['A2'].font = Font(italic=True)

        # Add game data
        game_df = pd.DataFrame(conf_game_data[threshold])
        for r_idx, row in enumerate(dataframe_to_rows(game_df, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 4:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif isinstance(value, (int, float)) and c_idx >= 8:  # PnL columns
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

    # Save workbook
    wb.save(excel_filename)
    print(f"[+] Excel file saved to: {excel_filename}")
    print(f"[+] Total sheets created: {len(wb.sheetnames)}")
    print(f"[+] EV thresholds: {len(ev_game_data)}")
    print(f"[+] Confidence thresholds: {len(conf_game_data)}")

    print("\n" + "="*80)

if __name__ == "__main__":
    main()
