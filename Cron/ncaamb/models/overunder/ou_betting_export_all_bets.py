#!/usr/bin/env python3
"""
Export ALL 2025 O/U bets - both Over and Under for each game
Shows what would have happened if you bet on either side
Uses O/U ensemble (XGBoost + LightGBM + CatBoost)
"""

import polars as pl
import numpy as np
import os
import pandas as pd
from datetime import datetime
from pathlib import Path
import sys

# Add parent directory to path
ncaamb_dir = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ncaamb_dir))

from models.overunder.ensemble3_model import Ensemble3Model


def load_features(years):
    """Load and combine features from multiple years"""
    dfs = []
    features_dir = ncaamb_dir

    first_loaded = False
    reference_schema = None

    for year in years:
        file = features_dir / f"features{year}.csv"
        if not file.exists():
            continue
        print(f"[+] Loading {file}...")
        df = pl.read_csv(file)

        if not first_loaded:
            for col in df.columns:
                if df[col].dtype in [pl.Float32, pl.Float64]:
                    df = df.with_columns(pl.col(col).cast(pl.Float64))
                elif df[col].dtype in [pl.Int32, pl.Int64]:
                    df = df.with_columns(pl.col(col).cast(pl.Int64))
            reference_schema = df.schema
            first_loaded = True
        else:
            for col in df.columns:
                if col in reference_schema:
                    target_dtype = reference_schema[col]
                    if df[col].dtype != target_dtype:
                        try:
                            df = df.with_columns(pl.col(col).cast(target_dtype))
                        except:
                            pass

        dfs.append(df)

    combined = pl.concat(dfs, how='diagonal')
    return combined


def create_ou_target_variable(df):
    """Create binary target variable for O/U"""
    df = df.with_columns([
        pl.when(
            (pl.col('actual_total').is_not_null()) &
            (pl.col('betonline_ou_line').is_not_null())
        ).then(
            pl.when(pl.col('actual_total') > pl.col('betonline_ou_line')).then(1).otherwise(0)
        ).otherwise(None).alias('ou_target')
    ])
    return df


def preprocess_features(df):
    """Convert odds columns to float"""
    odds_cols = [
        'betonline_ou_line', 'betonline_over_odds', 'betonline_under_odds',
        'avg_ou_line', 'avg_over_odds', 'avg_under_odds',
    ]
    for col in odds_cols:
        if col in df.columns:
            df = df.with_columns(pl.col(col).cast(pl.Float64))
    return df


def filter_quality_games(df):
    """Filter games with valid data"""
    initial_count = len(df)
    df = df.filter(
        (pl.col('team_1_data_quality') >= 0.5) &
        (pl.col('team_2_data_quality') >= 0.5)
    )
    df = df.filter(pl.col('ou_target').is_not_null())
    # At least one sportsbook needs O/U data
    df = df.filter(
        (pl.col('betonline_ou_line').is_not_null()) |
        (pl.col('bovada_ou_line').is_not_null()) |
        (pl.col('mybookie_ou_line').is_not_null())
    )
    removed = initial_count - len(df)
    print(f"[+] Filtered: {initial_count} -> {len(df)} ({removed} removed)")
    return df


def get_best_ou_odds(row, bet_type):
    """Get best odds for Over/Under, preferring best across all sportsbooks"""
    if bet_type == 'Over':
        odds_list = [
            (row.get('betonline_over_odds'), 'Betonline'),
            (row.get('bovada_over_odds'), 'Bovada'),
            (row.get('mybookie_over_odds'), 'MyBookie'),
        ]
    else:  # Under
        odds_list = [
            (row.get('betonline_under_odds'), 'Betonline'),
            (row.get('bovada_under_odds'), 'Bovada'),
            (row.get('mybookie_under_odds'), 'MyBookie'),
        ]

    best_odds = None
    best_book = None
    best_decimal = 0

    for odds, book in odds_list:
        if odds and odds != 0 and not np.isnan(odds):
            decimal = american_to_decimal(odds)
            if decimal > best_decimal:
                best_decimal = decimal
                best_odds = float(odds)
                best_book = book

    return best_odds, best_book


def american_to_decimal(american_odds):
    """Convert American odds to decimal"""
    american_odds = float(american_odds)
    if american_odds >= 0:
        return (american_odds / 100) + 1
    else:
        return (100 / abs(american_odds)) + 1


def calculate_ev(win_prob, decimal_odds):
    """Calculate Expected Value"""
    win_prob = float(win_prob)
    decimal_odds = float(decimal_odds)
    ev = (win_prob * (decimal_odds - 1)) - (1 - win_prob)
    return ev


def sigmoid(x, scale=3.0):
    """Convert point prediction to confidence probability"""
    return 1.0 / (1.0 + np.exp(-x / scale))


def main():
    print("="*80)
    print("EXPORT ALL 2025 O/U BETS - BOTH OVER/UNDER PER GAME")
    print("="*80)

    # Load data
    print("\n[STEP 1] Loading training data (2021-2024)...")
    train_df = load_features(['2021', '2022', '2023', '2024'])

    print("[STEP 2] Loading test data (2025)...")
    test_df = load_features(['2025'])

    # Prepare
    print("[STEP 3] Processing data...")
    train_df = create_ou_target_variable(train_df)
    test_df = create_ou_target_variable(test_df)
    train_df = preprocess_features(train_df)
    test_df = preprocess_features(test_df)

    print("[STEP 4] Filtering quality games...")
    train_df = filter_quality_games(train_df)
    test_df = filter_quality_games(test_df)

    print(f"[+] Train: {len(train_df)} games, Test: {len(test_df)} games")

    # Load and train ensemble
    print("\n[STEP 5] Training O/U ensemble (XGB + LGB + CatBoost)...")
    ensemble = Ensemble3Model()
    metrics = ensemble.train(train_df, test_size=0.2)

    print(f"[+] Ensemble Test MAE: {metrics['ensemble_test_mae']:.4f}")
    print(f"[+] Weights: XGB={metrics['xgb_weight']:.3f}, LGB={metrics['lgb_weight']:.3f}, Cat={metrics['cat_weight']:.3f}")

    # Get predictions on test data
    print("\n[STEP 6] Getting ensemble predictions on 2025 test data...")
    predictions = ensemble.predict(test_df)

    # Convert test_df to dict for easier access
    test_dict = test_df.to_dicts()
    predicted_totals = predictions['predicted_total']
    actual_totals = predictions['actual_total']

    print("\n[STEP 7] Creating betting records for all games...")
    all_bets = []

    for i, game in enumerate(test_dict):
        game_id = game.get('game_id', '')
        date = game.get('date', '')
        team_1 = game.get('team_1', '')
        team_2 = game.get('team_2', '')
        actual_total = actual_totals[i] if i < len(actual_totals) else None

        # Ensemble prediction
        predicted_total = predicted_totals[i]

        # Get O/U lines from all sportsbooks
        betonline_line = game.get('betonline_ou_line')
        bovada_line = game.get('bovada_ou_line')
        mybookie_line = game.get('mybookie_ou_line')

        # Skip if no O/U data available
        if all(x is None or np.isnan(x) for x in [betonline_line, bovada_line, mybookie_line]):
            continue

        # Convert point predictions to confidence (use betonline line as reference)
        ref_line = betonline_line if betonline_line and not np.isnan(betonline_line) else (bovada_line if bovada_line and not np.isnan(bovada_line) else mybookie_line)
        if ref_line is None or np.isnan(ref_line):
            continue

        xgb_pred = predictions['xgb_predicted_total'][i]
        lgb_pred = predictions['lgb_predicted_total'][i]
        cat_pred = predictions['cat_predicted_total'][i]

        xgb_conf_over = sigmoid(xgb_pred - ref_line, scale=3.0)
        lgb_conf_over = sigmoid(lgb_pred - ref_line, scale=3.0)
        cat_conf_over = sigmoid(cat_pred - ref_line, scale=3.0)

        # Ensemble confidence (weighted)
        ensemble_conf_over = (
            metrics['xgb_weight'] * xgb_conf_over +
            metrics['lgb_weight'] * lgb_conf_over +
            metrics['cat_weight'] * cat_conf_over
        )

        # Determine prediction direction
        if ensemble_conf_over > 0.5:
            confidence = ensemble_conf_over
            bet_dir = 'Over'
        else:
            confidence = 1.0 - ensemble_conf_over
            bet_dir = 'Under'

        # Get best odds across all sportsbooks
        if bet_dir == 'Over':
            odds, best_book = get_best_ou_odds(game, 'Over')
        else:
            odds, best_book = get_best_ou_odds(game, 'Under')

        if odds is None:
            continue

        decimal = american_to_decimal(odds)
        ev = calculate_ev(confidence, decimal)
        potential_profit = round(10 * (decimal - 1), 2)
        potential_loss = -10

        all_bets.append({
            'game_id': game_id,
            'date': date,
            'team_1': team_1,
            'team_2': team_2,
            'betonline_line': round(betonline_line, 1) if betonline_line and not np.isnan(betonline_line) else None,
            'bovada_line': round(bovada_line, 1) if bovada_line and not np.isnan(bovada_line) else None,
            'mybookie_line': round(mybookie_line, 1) if mybookie_line and not np.isnan(mybookie_line) else None,
            'predicted_total': round(predicted_total, 1),
            'actual_total': round(actual_total, 1) if actual_total and not np.isnan(actual_total) else None,
            'confidence': round(confidence, 4),
            'best_book': best_book,
            'odds_american': round(odds, 0),
            'odds_decimal': round(decimal, 3),
            'ev_%': round(ev * 100, 2),
            'potential_profit': potential_profit,
            'potential_loss': potential_loss,
        })

    # Export to Excel and CSV
    df_bets = pd.DataFrame(all_bets)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_dir = Path(__file__).parent / "saved"
    saved_dir.mkdir(parents=True, exist_ok=True)

    # Excel export with formatting
    excel_filename = saved_dir / f'ou_all_bets_2025_{timestamp}.xlsx'
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df_bets.to_excel(writer, sheet_name='OU_Bets', index=False)

        # Format the workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = writer.book
        worksheet = writer.sheets['OU_Bets']

        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color code wins (green) and losses (red)
        for row in worksheet.iter_rows(min_row=2, max_row=len(df_bets)+1):
            for cell in row:
                if cell.column == 14:  # PnL column (M)
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        widths = {
            'A': 18,  # game_id
            'B': 12,  # date
            'C': 20,  # team_1
            'D': 20,  # team_2
            'E': 10,  # ou_line
            'F': 15,  # predicted_total
            'G': 15,  # actual_total
            'H': 12,  # bet_type
            'I': 12,  # confidence
            'J': 12,  # sportsbook
            'K': 14,  # odds_american
            'L': 14,  # odds_decimal
            'M': 10,  # ev_%
            'N': 10,  # outcome
            'O': 10,  # pnl
        }
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width

    print(f"\n[+] Excel file saved: {excel_filename}")

    # CSV export
    csv_filename = saved_dir / f'ou_all_bets_2025_{timestamp}.csv'
    df_bets.to_csv(csv_filename, index=False)
    print(f"[+] CSV file saved: {csv_filename}")
    print(f"[+] Total bets exported: {len(all_bets)}")
    print(f"[+] Games: {len(test_dict)}")

    # Summary stats
    avg_confidence = np.mean([b['confidence'] for b in all_bets]) if all_bets else 0
    avg_ev = np.mean([b['ev_%'] for b in all_bets]) if all_bets else 0

    print(f"\n[SUMMARY - ALL O/U BETS]")
    print(f"  Total Bets: {len(all_bets)}")
    print(f"  Avg Confidence: {avg_confidence:.4f}")
    print(f"  Avg EV%: {avg_ev:.2f}%")
    print("="*80)


if __name__ == "__main__":
    main()
