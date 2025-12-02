#!/usr/bin/env python3
"""
Export 2025 O/U GOOD BETS - Filtered betting recommendations using Random Forest
Trains on 2021-2024 good bets classifier, tests on 2025
Only exports high-confidence Over/Under bets
"""

import os
import sys
from pathlib import Path
import polars as pl
import numpy as np
import pandas as pd
from datetime import datetime
import pickle
import warnings
warnings.filterwarnings('ignore')

# Add parent directory to path
ncaamb_dir = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ncaamb_dir))

from models.overunder.ensemble3_model import Ensemble3Model


def load_features_by_year(years: list):
    """Load feature files for specified years"""
    features_dir = ncaamb_dir
    all_features = []

    print(f"Loading features for years: {years}")
    for year in years:
        features_file = features_dir / f"features{year}.csv"
        if features_file.exists():
            print(f"  Loading features{year}.csv...")
            try:
                df = pd.read_csv(features_file)
                print(f"    [OK] Loaded {len(df)} games")
                all_features.append(df)
            except Exception as e:
                print(f"    [ERR] Error loading {year}: {e}")
        else:
            print(f"    [ERR] File not found: {features_file}")

    if not all_features:
        return None

    # Concatenate all years
    combined_df = pd.concat(all_features, ignore_index=True)

    # Convert numeric-looking columns to float
    for col in combined_df.columns:
        if any(x in col.lower() for x in ['odds', 'line', 'spread', 'total', 'score', 'pace', 'efg', 'adj', 'avg']):
            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')

    print(f"[OK] Combined: {len(combined_df)} total games\n")
    return combined_df


def create_ou_target_variable(df):
    """Create binary target variable for over/under betting"""
    # Filter: must have actual total and betonline line
    df_filtered = df.dropna(subset=['actual_total', 'betonline_ou_line'])

    before = len(df)
    after = len(df_filtered)
    print(f"Filtered for betonline O/U line: removed {before - after}, kept {after}")

    # Create target: 1 if over, 0 if under
    df_with_target = df_filtered.copy()
    df_with_target['ou_target'] = (df_with_target['actual_total'] > df_with_target['betonline_ou_line']).astype(int)

    over_count = (df_with_target['ou_target'] == 1).sum()
    under_count = (df_with_target['ou_target'] == 0).sum()

    print(f"Target distribution:")
    print(f"  Over hits (1):  {over_count}")
    print(f"  Under hits (0): {under_count}\n")

    return df_with_target


def sigmoid(x, scale=3.0):
    """Convert point prediction to confidence probability"""
    return 1.0 / (1.0 + np.exp(-np.clip(x, -500, 500) / scale))


def get_predictions_all_models(df, ensemble_model) -> tuple:
    """Get predictions from all 3 ensemble models"""
    # Get numeric columns that exist in df
    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
    # Exclude target and metadata
    exclude = {'actual_total', 'team_1_score', 'team_2_score', 'game_id', 'date', 'team_1', 'team_2', 'ou_target', 'betonline_ou_line'}
    feature_cols = [c for c in numeric_cols if c not in exclude]

    X = df[feature_cols].fillna(0).values

    # Convert to Polars for ensemble prediction
    df_pl = pl.from_pandas(df)
    predictions = ensemble_model.predict(df_pl)

    xgb_preds = np.array(predictions['xgb_predicted_total'])
    lgb_preds = np.array(predictions['lgb_predicted_total'])
    cat_preds = np.array(predictions['cat_predicted_total'])

    print(f"[OK] Generated predictions from all 3 models\n")
    return xgb_preds, lgb_preds, cat_preds


def create_ou_good_bets_data(df, ensemble_model, xgb_preds: np.ndarray,
                             lgb_preds: np.ndarray, cat_preds: np.ndarray) -> tuple:
    """
    Create good bets training data for O/U using Random Forest
    Features based on model predictions and odds
    Uses betonline_ou_line for feature calculation
    """
    # Filter to rows with ou_target and betonline line
    df_filtered = df.dropna(subset=['ou_target', 'betonline_ou_line']).copy()

    # Get ensemble weights from model
    xgb_weight = ensemble_model.xgb_weight
    lgb_weight = ensemble_model.lgb_weight
    cat_weight = ensemble_model.cat_weight

    # Add predictions
    df_filtered['xgb_point_pred'] = xgb_preds[:len(df_filtered)]
    df_filtered['lgb_point_pred'] = lgb_preds[:len(df_filtered)]
    df_filtered['cat_point_pred'] = cat_preds[:len(df_filtered)]

    # Convert point predictions to confidence using sigmoid (using betonline line for feature calculation)
    df_filtered['xgb_confidence_over'] = 1.0 / (1.0 + np.exp(-np.clip(df_filtered['xgb_point_pred'] - df_filtered['betonline_ou_line'], -500, 500) / 3.0))
    df_filtered['lgb_confidence_over'] = 1.0 / (1.0 + np.exp(-np.clip(df_filtered['lgb_point_pred'] - df_filtered['betonline_ou_line'], -500, 500) / 3.0))
    df_filtered['cat_confidence_over'] = 1.0 / (1.0 + np.exp(-np.clip(df_filtered['cat_point_pred'] - df_filtered['betonline_ou_line'], -500, 500) / 3.0))

    # Clip to [0, 1]
    for col in ['xgb_confidence_over', 'lgb_confidence_over', 'cat_confidence_over']:
        df_filtered[col] = df_filtered[col].clip(0.0, 1.0)

    # Ensemble confidence (weighted average)
    df_filtered['ensemble_confidence_over'] = (
        xgb_weight * df_filtered['xgb_confidence_over'] +
        lgb_weight * df_filtered['lgb_confidence_over'] +
        cat_weight * df_filtered['cat_confidence_over']
    )

    # Model disagreement: std dev of 3 confidences
    confs = np.array([
        df_filtered['xgb_confidence_over'].values,
        df_filtered['lgb_confidence_over'].values,
        df_filtered['cat_confidence_over'].values
    ])
    df_filtered['model_std_dev'] = np.std(confs, axis=0)

    # Fill missing odds data
    for col in ['betonline_ou_line', 'avg_ou_line', 'ou_line_variance',
                'avg_over_odds', 'avg_under_odds',
                'betonline_over_odds', 'betonline_under_odds']:
        if col not in df_filtered.columns:
            df_filtered[col] = 0.0
        else:
            df_filtered[col] = df_filtered[col].fillna(0.0)

    # Select feature columns
    feature_cols = [
        'xgb_confidence_over', 'lgb_confidence_over', 'cat_confidence_over',
        'ensemble_confidence_over', 'model_std_dev',
        'betonline_ou_line', 'avg_ou_line', 'ou_line_variance',
        'avg_over_odds', 'avg_under_odds',
        'betonline_over_odds', 'betonline_under_odds',
    ]

    selected_cols = [col for col in feature_cols if col in df_filtered.columns]
    X_bets = df_filtered[selected_cols].values
    y_bets = df_filtered['ou_target'].values

    print(f"Created {len(X_bets)} betting rows ({len(df)} games)")
    print(f"  Over hits (1): {int(np.sum(y_bets))}")
    print(f"  Under hits (0): {int(len(y_bets) - np.sum(y_bets))}\n")

    return X_bets, y_bets, selected_cols, df_filtered


def american_to_decimal(american_odds):
    """Convert American odds to decimal"""
    american_odds = float(american_odds)
    if american_odds >= 0:
        return (american_odds / 100) + 1
    else:
        return (100 / abs(american_odds)) + 1


def calculate_ev(win_prob, decimal_odds):
    """Calculate Expected Value"""
    win_prob = float(win_prob)
    decimal_odds = float(decimal_odds)
    ev = (win_prob * (decimal_odds - 1)) - (1 - win_prob)
    return ev


def get_best_ou_odds(row, bet_type):
    """Get best odds for Over/Under"""
    if bet_type == 'Over':
        if row.get('betonline_over_odds') and row.get('betonline_over_odds') != 0:
            return float(row['betonline_over_odds']), 'Betonline'
        if row.get('avg_over_odds') and row.get('avg_over_odds') != 0:
            return float(row['avg_over_odds']), 'Average'
    else:  # Under
        if row.get('betonline_under_odds') and row.get('betonline_under_odds') != 0:
            return float(row['betonline_under_odds']), 'Betonline'
        if row.get('avg_under_odds') and row.get('avg_under_odds') != 0:
            return float(row['avg_under_odds']), 'Average'

    return None, None


def main():
    print("\n")
    print("="*80)
    print("O/U GOOD BETS EXPORT - FILTERED BETTING RECOMMENDATIONS")
    print("Using O/U ensemble + Random Forest good bets classifier")
    print("Training on 2021-2024, Testing on 2025")
    print("="*80 + "\n")

    # Load training data
    print("STEP 1: Loading Training Data (2021-2024)")
    print("-"*80 + "\n")
    train_df = load_features_by_year(['2021', '2022', '2023', '2024'])

    if train_df is None or len(train_df) == 0:
        print("Failed to load training features")
        return

    print("STEP 2: Creating Target Variable")
    print("-"*80 + "\n")
    train_df = create_ou_target_variable(train_df)

    if len(train_df) == 0:
        print("No training games with betonline O/U data")
        return

    # Train ensemble on training data
    print("STEP 3: Training O/U Ensemble (2021-2024)")
    print("-"*80 + "\n")
    ensemble = Ensemble3Model()
    train_df_pl = pl.from_pandas(train_df)
    metrics = ensemble.train(train_df_pl, test_size=0.2)
    print(f"[+] Ensemble Test MAE: {metrics['ensemble_test_mae']:.4f}")
    print(f"[+] Weights: XGB={metrics['xgb_weight']:.3f}, LGB={metrics['lgb_weight']:.3f}, Cat={metrics['cat_weight']:.3f}\n")

    # Get predictions on training data
    print("STEP 4: Getting Ensemble Predictions (Training Data)")
    print("-"*80 + "\n")
    xgb_preds_train, lgb_preds_train, cat_preds_train = get_predictions_all_models(
        train_df, ensemble
    )

    # Create good bets data
    print("STEP 5: Creating Good Bets Training Data")
    print("-"*80 + "\n")
    X_bets_train, y_bets_train, gb_feature_cols, train_df_bets = create_ou_good_bets_data(
        train_df, ensemble, xgb_preds_train, lgb_preds_train, cat_preds_train
    )

    # Train Random Forest
    print("STEP 6: Training Random Forest Good Bets Model")
    print("-"*80 + "\n")

    from sklearn.ensemble import RandomForestClassifier

    print("Training Random Forest on good bets data...")
    rf_model = RandomForestClassifier(
        n_estimators=100,
        max_depth=10,
        min_samples_split=50,
        min_samples_leaf=20,
        random_state=42,
        n_jobs=-1,
        verbose=0
    )

    rf_model.fit(X_bets_train, y_bets_train)
    print("[OK] Random Forest training complete\n")

    # Load test data
    print("STEP 7: Loading Test Data (2025)")
    print("-"*80 + "\n")

    test_df = load_features_by_year(['2025'])

    if test_df is None or len(test_df) == 0:
        print("No test data (2025) available")
        return

    print("STEP 8: Creating Test Target Variable")
    print("-"*80 + "\n")
    test_df = create_ou_target_variable(test_df)

    if len(test_df) == 0:
        print("No test games with betonline O/U data")
        return

    # Get predictions on test set
    print("STEP 9: Getting Ensemble Predictions (Test Data)")
    print("-"*80 + "\n")
    xgb_preds_test, lgb_preds_test, cat_preds_test = get_predictions_all_models(
        test_df, ensemble
    )

    # Create good bets data for test set
    print("STEP 10: Creating Good Bets Test Data")
    print("-"*80 + "\n")
    X_bets_test, y_bets_test, _, test_df_bets = create_ou_good_bets_data(
        test_df, ensemble, xgb_preds_test, lgb_preds_test, cat_preds_test
    )

    # Get good bets predictions and probabilities
    print("STEP 11: Classifying Good Bets")
    print("-"*80 + "\n")
    y_pred_test = rf_model.predict(X_bets_test)
    y_proba_test = rf_model.predict_proba(X_bets_test)[:, 1]  # Probability of "good bet" (over hit)

    # Filter for good bets
    good_bets_mask = (y_proba_test >= 0.5)  # High confidence good bets
    good_bets_indices = np.where(good_bets_mask)[0]

    print(f"[+] Total games analyzed: {len(test_df_bets)}")
    print(f"[+] Good bets found: {len(good_bets_indices)}\n")

    # Create export records
    print("STEP 12: Creating Export Records")
    print("-"*80 + "\n")

    all_bets = []
    for idx in good_bets_indices:
        game_row = test_df_bets.iloc[idx]
        game_id = game_row.get('game_id', '')
        date = game_row.get('date', '')
        team_1 = game_row.get('team_1', '')
        team_2 = game_row.get('team_2', '')
        actual_total = game_row.get('actual_total')
        good_bet_prob = y_proba_test[idx]

        ensemble_conf_over = game_row.get('ensemble_confidence_over', 0)
        ensemble_conf_under = 1.0 - ensemble_conf_over

        # Get O/U lines from all sportsbooks
        betonline_line = game_row.get('betonline_ou_line')
        bovada_line = game_row.get('bovada_ou_line')
        mybookie_line = game_row.get('mybookie_ou_line')

        # Determine bet recommendation based on confidence
        if ensemble_conf_over > 0.55:
            confidence = ensemble_conf_over
            odds_over = game_row.get('betonline_over_odds') or game_row.get('bovada_over_odds') or game_row.get('mybookie_over_odds')
            odds_under = None
        elif ensemble_conf_under > 0.55:
            confidence = ensemble_conf_under
            odds_under = game_row.get('betonline_under_odds') or game_row.get('bovada_under_odds') or game_row.get('mybookie_under_odds')
            odds_over = None
        else:
            # Skip if no strong recommendation
            continue

        # Get best odds
        odds = odds_over if odds_over else odds_under
        if odds is None or np.isnan(odds):
            continue

        decimal = american_to_decimal(odds)
        ev = calculate_ev(confidence, decimal)
        potential_profit = round(10 * (decimal - 1), 2)
        potential_loss = -10

        all_bets.append({
            'game_id': game_id,
            'date': date,
            'team_1': team_1,
            'team_2': team_2,
            'betonline_line': round(betonline_line, 1) if betonline_line and not np.isnan(betonline_line) else None,
            'bovada_line': round(bovada_line, 1) if bovada_line and not np.isnan(bovada_line) else None,
            'mybookie_line': round(mybookie_line, 1) if mybookie_line and not np.isnan(mybookie_line) else None,
            'actual_total': round(actual_total, 1) if actual_total and not np.isnan(actual_total) else None,
            'confidence': round(confidence, 4),
            'good_bet_score': round(good_bet_prob, 4),
            'odds_american': round(odds, 0),
            'odds_decimal': round(decimal, 3),
            'ev_%': round(ev * 100, 2),
            'potential_profit': potential_profit,
            'potential_loss': potential_loss,
        })

    if not all_bets:
        print("No high-confidence good bets found for 2025")
        return

    # Export to Excel and CSV
    df_bets = pd.DataFrame(all_bets)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_dir = Path(__file__).parent / "saved"
    saved_dir.mkdir(parents=True, exist_ok=True)

    # Excel export with formatting
    excel_filename = saved_dir / f'ou_good_bets_2025_{timestamp}.xlsx'
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df_bets.to_excel(writer, sheet_name='Good_Bets', index=False)

        # Format the workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = writer.book
        worksheet = writer.sheets['Good_Bets']

        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color code wins (green) and losses (red)
        for row in worksheet.iter_rows(min_row=2, max_row=len(df_bets)+1):
            for cell in row:
                if cell.column == 15:  # PnL column
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        widths = {
            'A': 18, 'B': 12, 'C': 20, 'D': 20, 'E': 10,
            'F': 15, 'G': 15, 'H': 18, 'I': 15, 'J': 15,
            'K': 12, 'L': 14, 'M': 14, 'N': 10, 'O': 10,
        }
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width

    print(f"\n[+] Excel file saved: {excel_filename}")

    # CSV export
    csv_filename = saved_dir / f'ou_good_bets_2025_{timestamp}.csv'
    df_bets.to_csv(csv_filename, index=False)
    print(f"[+] CSV file saved: {csv_filename}")
    print(f"[+] Good bets exported: {len(all_bets)}")

    # Summary stats
    avg_confidence = np.mean([b['confidence'] for b in all_bets]) if all_bets else 0
    avg_good_bet_score = np.mean([b['good_bet_score'] for b in all_bets]) if all_bets else 0
    avg_ev = np.mean([b['ev_%'] for b in all_bets]) if all_bets else 0

    print(f"\n[SUMMARY - O/U GOOD BETS]")
    print(f"  Total Bets: {len(all_bets)}")
    print(f"  Avg Confidence: {avg_confidence:.4f}")
    print(f"  Avg Good Bet Score: {avg_good_bet_score:.4f}")
    print(f"  Avg EV%: {avg_ev:.2f}%")
    print("="*80)


if __name__ == "__main__":
    main()
